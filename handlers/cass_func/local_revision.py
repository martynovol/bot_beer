import asyncio
import json,requests

from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram import types, Dispatcher
from aiogram.dispatcher.filters import Text
from aiogram.types import CallbackQuery, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, InlineKeyboardMarkup, \
    InlineKeyboardButton
import sqlite3 as sq
from create_bot import dp, bot

import os
from id import token
from database import sqlite_db

from keyboards import admin_kb, admin_cancel_kb, cassier_kb, mod_kb, kb_client, main_cassier_kb

from datetime import datetime
from datetime import date, timedelta

import gspread
from gspread import Cell, Client, Spreadsheet, Worksheet
from gspread.utils import rowcol_to_a1

from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import openpyxl

from handlers import inf
from handlers import emoji_bot
from handlers.mod_func import action_with_user

import emoji

get_data = {}

def global_dictionary(user_id, method="check", data=None):
    if method == "add":
        if user_id in get_data:
            get_data[user_id][data[0]] = [data[1], False]
        else:
            get_data[user_id] = {}
            get_data[user_id][data[0]] = [data[1], False]
    elif method == "check":
        if data == None:
            for product in get_data[user_id].keys():
                if get_data[user_id][product][1] == False:
                    return product, get_data[user_id][product][0]
            return 0, 0
        else:
            return get_data[user_id]
    elif method == "turn":
        get_data[user_id][data][1] = True
    elif method == "clear":
        get_data[user_id] = {}

class FSMCount(StatesGroup):
    product_count = State()
    agree_count = State()


class FSMTable_stocks(StatesGroup):
    date1 = State()


class FSMGet_stock(StatesGroup):
    date = State()
    point = State()

def check_now_open(user_id):
    now_point, now_date = "", ""
    for ret in sqlite_db.cur.execute('SELECT now_user, point1, date1 FROM reports_open WHERE report_close_id = ?',
                                        ['']).fetchall():
        if str(inf.get_user_id(ret[0])) == str(user_id):
            now_point, now_date = ret[1], ret[2]
    return now_point, now_date

async def cm_start_position(message: types.Message):
    now_point, now_date = check_now_open(message.from_user.id)
    if now_point == "" or now_date == "":
        await bot.send_message(message.from_user.id, f'Смена не была открыта')
        return
    
    check_revision = False
    day, month, year = now_date.split('.')
    for ret in sqlite_db.cur.execute("SELECT * FROM local_revision WHERE day = ? AND month = ? AND year = ? AND point_id = ?", [day, month, year, inf.get_id_point(now_point)]):
        check_revision = ret[0]
    if check_revision != False:
        await bot.send_message(message.from_user.id, "Ревизия за сегодня уже была проведена.")
        return

    mes = await bot.send_message(message.from_user.id, "Подождите идёт выгрузка...")
    now_product = get_retail_oborots_for_day(now_date, now_point)
    global_dictionary(message.from_user.id, "clear")
    for name, href in now_product.items():
        global_dictionary(message.from_user.id, "add", [name, href])
    await bot.delete_message(message.from_user.id, mes.message_id)
    mes = await bot.send_message(message.from_user.id, f"Впишите количество {list(now_product.keys())[0]}, которые находятся на точке в момент загрузки:", reply_markup=ReplyKeyboardMarkup(resize_keyboard=True).add("Отмена"))
    await FSMCount.product_count.set()


async def state_get_count(message: types.Message, state: FSMContext):
    now_point, now_date = check_now_open(message.from_user.id)
    if now_point == "" or now_date == "" or not message.text.isdigit():
        return
    async with state.proxy() as data:
        print(data)
    
    name_product, now_product = global_dictionary(message.from_user.id)
    
    if now_product != None:
        stock = get_now_stock_product(now_product, inf.get_id_sklad_point(now_point))
        if stock == int(message.text):
            async with state.proxy() as data:
                data[name_product] = [str(stock), message.text]
            global_dictionary(message.from_user.id, "turn", name_product)
            next_product = global_dictionary(message.from_user.id)[0]
            if next_product != 0:
                await bot.send_message(message.from_user.id, f"Впишите количество {next_product} на точке:")
                await FSMCount.product_count.set()
            else:
                text = f"Локальная ревизия завершена, данные отправлены."
                #text += "\nЕсли вы отправили не те данные, пройдите локальную ревизию ещё раз - данные будут перезаписаны"
                await bot.send_message(message.from_user.id, f"{text}.", reply_markup=inf.kb(message.from_user.id))
                await sqlite_db.sql_add_local_revision(state, message.from_user.id, inf.get_id_point(now_point), now_date)
                await state.finish()

        else:
            await bot.send_message(message.from_user.id, f"⚠️Данные не сходятся. Перепроверьте и подтвердите, вписав количество ещё раз:")
            await FSMCount.agree_count.set()


async def agree_get_count(message: types.Message, state: FSMContext):
    now_point, now_date = check_now_open(message.from_user.id)
    if now_point == "" or now_date == "" or not message.text.isdigit():
        return
    
    name_product, now_product = global_dictionary(message.from_user.id)

    stock = get_now_stock_product(now_product, inf.get_id_sklad_point(now_point))
    async with state.proxy() as data:
        data[name_product] = [stock, message.text]
    global_dictionary(message.from_user.id, "turn", name_product)
    next_product = global_dictionary(message.from_user.id)[0]
    if next_product != 0:
        await bot.send_message(message.from_user.id, f"Впишите количество {next_product} на точке:")
        await FSMCount.product_count.set()
    else:
        text = f"Локальная ревизия завершена, данные отправлены."
        #text += "\nЕсли вы отправили не те данные, пройдите локальную ревизию ещё раз - данные будут перезаписаны"
        await bot.send_message(message.from_user.id, f"{text}.", reply_markup=inf.kb(message.from_user.id))
        await sqlite_db.sql_add_local_revision(state, message.from_user.id, inf.get_id_point(now_point), now_date)
        await state.finish()




def get_now_stock_product(product_href, url_restore):
    url_store = get_store(f"https://api.moysklad.ru/api/remap/1.2/entity/retailstore/{url_restore}")
    url_get = f"https://api.moysklad.ru/api/remap/1.2/report/stock/bystore?filter=store={url_store};product={product_href}"
    response = requests.get(url_get, headers=token.headers)
    data = json.loads(response.text)
    
    if len(data['rows']) != 0:
        return int(data['rows'][0]['stockByStore'][0]['stock'])
    else:
        return 0


def get_store(url_store):
    response = requests.get(url_store, headers=token.headers)
    data = json.loads(response.text)
    return data['store']['meta']['href']


def get_retail_oborots_for_day(now_date, now_point):
    momentFrom, momentTo = get_moments(now_date)
    retail_store_id = f"https://api.moysklad.ru/api/remap/1.2/entity/retailstore/{inf.get_id_sklad_point(now_point)}"
    url_get = f"https://api.moysklad.ru/api/remap/1.2/report/turnover/all?filter=retailStore={retail_store_id}"
    response = requests.get(url_get, headers=token.headers, params={'momentFrom': momentFrom, "momentTo": momentTo,})
    data = json.loads(response.text)
    result_report = {}
    try:
        for obj in data['rows']:
            result_report[obj["assortment"]['name']] = obj["assortment"]['meta']['href']
    except Exception:
        print(data)
    return result_report
    


def get_store(url_store):
    response = requests.get(url_store, headers=token.headers)
    data = json.loads(response.text)
    return data['store']['meta']['href']


def get_moments(now_date):
    normal_date = f"{now_date.split('.')[2]}-{now_date.split('.')[1]}-{now_date.split('.')[0]}"
    last_day = datetime.strptime(normal_date, "%Y-%m-%d") - timedelta(days=1)
    normal_last_date = f"{datetime.strftime(last_day, '%Y-%m-%d')} 00:00:00"
    normal_date = normal_date + " 02:00:00"
    return normal_last_date, normal_date




async def cm_start_get_revision(message: types.Message):
    admins = inf.get_admin_id()
    if message.from_user.id not in admins:
        await bot.send_message(message.from_user.id, 'Эта функция доступна только администраторам')
        return
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(datetime.now().strftime('%d.%m.%Y'))
    kb.row((datetime.now() - timedelta(days=1)).strftime('%d.%m.%Y'))
    kb.row("Отмена")
    await bot.send_message(message.from_user.id, f"Введите дату проведения в формате ДД.ММ.ГГГГ:", reply_markup=kb)
    await FSMGet_stock.date.set()


async def get_stock_date(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['date'] = message.text
    stores = inf.get_name_shops()
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    for store in stores:
        kb.add(store)
    kb.row("Отмена")
    await bot.send_message(message.from_user.id, f"Введите точку проведения:", reply_markup=kb)
    await FSMGet_stock.next()



async def get_stock_point(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        day, month, year = data['date'].split('.')
        point = message.text
    text = f"Ревизия за {day}.{month}.{year} на {point}\n"
    for ret in sqlite_db.cur.execute("SELECT * FROM local_revision WHERE day = ? AND month = ? AND year = ? AND point_id = ?", [day, month, year, inf.get_id_point(point)]):
        id_revision = ret[0]
        text += f"Провёл: {inf.get_name(ret[1])} | Продавец {inf.get_name(ret[6])}\n\n"
        stock, revision = 0, 0
        for jet in sqlite_db.cur.execute("SELECT * FROM local_revision_position WHERE id_revision = ?", [id_revision]):
            text += f"{jet[2]}: ({jet[3]}/{jet[4]})\n"
            stock += int(jet[3])
            revision += int(jet[4])
        text += f"\nОстаток в МС: {stock} | Фактический остаток: {revision}\nИтог: {revision - stock}"
    await bot.send_message(message.from_user.id, text, reply_markup=inf.kb(message.from_user.id))
    await state.finish()



async def start_table_stocks(message: types.Message, state: FSMContext):
    admins = inf.get_admin_id()
    if message.from_user.id not in admins:
        await bot.send_message(message.from_user.id, 'Эта функция доступна только администраторам')
        return
    b11 = KeyboardButton(str(datetime.now().strftime('%m.%Y')))
    b12 = KeyboardButton('Назад')
    button_case_admin_report = ReplyKeyboardMarkup(resize_keyboard=True).add(b11).row(b12)
    await bot.send_message(message.from_user.id,
                           f'Впишите месяц и год за который нужно загрузить данные в таблицу в формате ММ.ГГГГ:',
                           reply_markup=button_case_admin_report)
    await FSMTable_stocks.date1.set()


async def get_stocks_finish(message: types.Message, state: FSMContext):
    await sqlite_db.info_admin(message.from_user.id, f"Выгружает таблицу")
    bit = message.text.split('.')
    if len(bit) != 2 or (len(bit[0]) != 1 and len(bit[0]) != 2) or len(bit[1]) != 4 or not bit[0].isdigit() or not bit[
        1].isdigit():
        await message.reply(f'Вы ввели некорректное число, попробуйте снова, формат "ММ.ГГГГ"')
        return
    month, year = bit[0], bit[1]
    await bot.send_message(message.from_user.id, f'Подождите, данные выгружаются',
                           reply_markup=inf.kb(message.from_user.id))
    
    await in_file_stocks(message.from_user.id, month, year)

    with open(f'Ревизии_{message.from_user.id}_{month}_{year}.xlsx', 'rb') as file:
        await bot.send_document(message.from_user.id, file)
    try:
        #in_file_stocks(message.from_user.id, month, year)
        pass
    except Exception:
        await bot.send_message(message.from_user.id, f'Произошла ошибка. Невозможно выгрузить таблицу',
                           reply_markup=inf.kb(message.from_user.id))
        await state.finish()
    
    try:
        os.remove(f'Ревизии_{message.from_user.id}_{month}_{year}.xlsx')
    except Exception:
        pass

    await state.finish()


def get_data_store(month, year, store):
    day_stocks = [["Дата проведения", "По складу", "Ревизия", "Разница", "Проверяющий", "Продавец"]]
    m_stock, m_now_stock = 0, 0
    uni_pos, all_pos = {}, []

    for ret in sqlite_db.cur.execute("SELECT id, user_id, point_id, day, month, year, last_user_id FROM local_revision WHERE month = ? AND year = ? AND point_id = ?", [month, year, inf.get_id_point(store)]).fetchall():
        total_stock, total_now_stock = 0, 0
        date_stock = f"{ret[3]}.{ret[4]}.{ret[5]}"
        
        for jet in sqlite_db.cur.execute("SELECT stock, now_stock, name_position FROM local_revision_position WHERE id_revision = ?", [ret[0]]).fetchall():
            total_stock += int(jet[0])
            total_now_stock += int(jet[1])
            uni_pos[jet[2]] = (int(jet[0]), int(jet[1]))
            all_pos.append([date_stock, jet[2], int(jet[0]), int(jet[1]), int(jet[1]) - int(jet[0])])
            

        m_stock += total_stock
        m_now_stock += total_now_stock

        day_stocks.append([date_stock, total_stock, total_now_stock, total_now_stock - total_stock, inf.get_name(ret[1]), inf.get_name(ret[6])])

    day_stocks.append([""]*6)

    stock_pos, rev_pos = 0, 0
    for position in uni_pos.values():
        stock_pos += position[0]
        rev_pos += position[1]

    day_stocks.append(["Уникальные позиции:", len(uni_pos)])
    day_stocks.append(["Остаток по МС:", stock_pos])
    day_stocks.append(["Ревизии:", rev_pos])
    day_stocks.append(["Разница:", rev_pos - stock_pos])

    day_stocks.append([""]*6)

    day_stocks.append(["Дата","Позиция", "Остаток по МС", "Ревизии", "Разница"])

    for value in all_pos:
        day_stocks.append(value)

    return day_stocks


async def in_file_stocks(uid, month, year):
    wb = openpyxl.Workbook()
    for store_name in inf.get_name_shops():
        data_store = get_data_store(month, year, store_name)
        wb.create_sheet(store_name)
        wb.active = wb[store_name]
        ws = wb.active

        for i, _str in enumerate(data_store):
            ws.append(_str)
            row = ws.row_dimensions[i + 1]
            bold_str = ["Уникальные позиции:", "Остаток по МС:", "Ревизии:", "Разница:", "Дата"]
            if _str[0] == "Дата":
                start_filter = i + 1
            if _str[0] in bold_str:
                row.font = Font(size = 13, bold=True)
            else:
                row.font = Font(size = 13)
            row.alignment = Alignment(horizontal='center')
        ws.auto_filter.ref = f'A{start_filter}:E{i + 1}'

        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 44
        ws.column_dimensions['C'].width = 17
        ws.column_dimensions['D'].width = 17
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30

        row = ws.row_dimensions[1]
        row.font = Font(bold=True, size=13)

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    wb.save(f'Ревизии_{uid}_{month}_{year}.xlsx')
    try:
        #wb.save(f'Ревизии_{uid}_{month}_{year}.xlsx')
        pass
    except Exception:
        os.remove(f'Ревизии_{uid}_{month}_{year}.xlsx')
        wb.save(f'Ревизии_{uid}_{month}_{year}.xlsx')


def register_handlers_local_report(dp: Dispatcher):
    dp.register_message_handler(cm_start_position,
                                Text(equals=f'📒Локальная ревизия', ignore_case=True), state=None)
    #dp.register_callback_query_handler(finish_report_open, lambda x: x.data and x.data.startswith('finish_open '))
    # dp.register_message_handler(add_photo_open, content_types=['any'], state=FSMProblem.problems)
    dp.register_message_handler(state_get_count, state=FSMCount.product_count)
    dp.register_message_handler(agree_get_count, state=FSMCount.agree_count)
    dp.register_message_handler(cm_start_get_revision,
                                Text(equals=f'📒Получить локальную ревизию', ignore_case=True), state=None)
    dp.register_message_handler(get_stock_date, state=FSMGet_stock.date)
    dp.register_message_handler(get_stock_point, state=FSMGet_stock.point)

    dp.register_message_handler(start_table_stocks, Text(equals=f'Выгрузить таблицу', ignore_case=True))
    dp.register_message_handler(get_stocks_finish, state=FSMTable_stocks.date1)