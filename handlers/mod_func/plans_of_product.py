import asyncio
import calendar
import os

from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram import types, Dispatcher
from aiogram.dispatcher.filters import Text
from aiogram.types import CallbackQuery, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, InlineKeyboardMarkup, \
    InlineKeyboardButton
import sqlite3 as sq
from create_bot import dp, bot
import requests, json
import openpyxl
from database import sqlite_db
from id import token
from keyboards import admin_kb, admin_cancel_kb, cassier_kb, mod_kb, kb_client, main_cassier_kb

from datetime import datetime
from datetime import date, timedelta

from handlers import inf
from handlers import emoji_bot


async def cm_start_plans_product(message: types.Message):
    await bot.send_message(message.from_user.id, "Пожалуйста, введите дату в формате 'ММ.ГГГГ' (например, '04.2024'):")
    await bot.register_next_step_handler(message, process_date_step_product)


async def process_date_step_product(message: types.Message):
    date = message.text
    await bot.send_message(message.from_user.id, "Теперь загрузите Excel-файл с учётом продаж:")
    await bot.register_next_step_handler(message, process_excel_step, date)


async def process_excel_step(message: types.Message, date: str):
    if message.content_type == 'document':
        file_id = message.document.file_id
        file_info = await bot.get_file(file_id)
        downloaded_file = await bot.download_file(file_info.file_path)
        wb = openpyxl.load_workbook(downloaded_file)
        sheet = wb.active
        sales_data = {}
        for column in range(1, sheet.max_column + 1, 3):
            shop_name = sheet.cell(row=1, column=column).value
            if shop_name:
                sales_data[shop_name] = {}
                for row in range(2, sheet.max_row + 1):
                    product_group = sheet.cell(row=row, column=column).value
                    if product_group:
                        sales_data[shop_name][product_group] = sheet.cell(row=row, column=column + 1).value
        
        
        print("Дата:", date)
        print("Учёт продаж:")
        print(sales_data)
        await bot.send_message(message.from_user.id, "Данные успешно обработаны и выведены в консоль.")
    else:
        await bot.send_message(message.from_user.id, "Пожалуйста, загрузите только Excel-файл.")


def register_handlers_plans_products(dp: Dispatcher):
    dp.register_message_handler(cm_start_plans_product, Text(equals=f'📔Продажи по товарам', ignore_case=True),state=None)
