import asyncio
import calendar
import os

from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram import types, Dispatcher
from aiogram.dispatcher.filters import Text
from aiogram.types import CallbackQuery, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, InlineKeyboardMarkup, \
    InlineKeyboardButton
import sqlite3 as sq
from create_bot import dp, bot
import requests, json
import openpyxl
from database import sqlite_db
from id import token
from keyboards import admin_kb, admin_cancel_kb, cassier_kb, mod_kb, kb_client, main_cassier_kb

from datetime import datetime
from datetime import date, timedelta

from handlers import inf
from handlers import emoji_bot


UPLOADS_DIR = "uploads"


class FSMTakePlans(StatesGroup):
    date = State()


class FSMPlans(StatesGroup):
    date = State()
    id = State()


class FSMGetPlans(StatesGroup):
    date = State()
    point = State()


async def cm_start_plans(message: types.Message):
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.insert(datetime.now().strftime('%m.%Y')).row('Отмена')
    await bot.send_message(message.from_user.id, f'Введите месяц и год для плана продаж:', reply_markup=kb)
    await FSMPlans.date.set()


async def plans_take_date(message: types.Message,  state: FSMContext):
    try:
        date_ = datetime.strptime(message.text, "%m.%Y").strftime("%m.%Y")
    except Exception:
        await bot.send_message(message.from_user.id, f'Формат(ММ.ГГГГ)')
        return

    async with state.proxy() as data:
        data['date'] = date_

    kb = ReplyKeyboardMarkup(resize_keyboard=True).add('Отмена')
    await get_pattern()
    await bot.send_message(message.from_user.id, f'Отправьте документ .xlsx с шаблоном плана продаж', reply_markup=kb)
    with open(f'pattern.xlsx', 'rb') as file:
        await bot.send_document(message.from_user.id, file)
    await FSMPlans.next()


async def plans_get(message: types.Message,  state: FSMContext):
    file_id = message.document.file_id

    file_path = os.path.join(UPLOADS_DIR, f"plans_of_sales.xlsx")
    await bot.download_file_by_id(file_id, file_path)
    plans = await read_excel_file(file_path, message.from_user.id)
    if None in plans.keys():
        await bot.send_message(message.from_user.id, f'Не удалось найти компонент: {plans[None]}')
        return
    async with state.proxy() as data:
        await sqlite_db.sql_add_plan(plans, data['date'])
    await message.reply(f'Файл обработан, планы продаж поставлены', reply_markup=inf.kb(message.from_user.id))
    await state.finish()


async def read_excel_file(file_path, uid):
    store_names = inf.get_name_shops()
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    result = {}
    all_groups = await get_name_groups()

    for y in range(1, 1000, 4):
        sh = ws.cell(1, y).value
        if sh in store_names:
            result[sh] = {}
            for x in range(3, 1000):
                group = ws.cell(x, y).value
                if group in all_groups.keys():

                    path_len = all_groups[group][0]
                    for path in all_groups[group]:

                        if len(path[1]) < len(path_len[1]):
                            path_len = path

                    result[sh][(group, path_len[0])] = \
                        [float(ws.cell(x, y + 1).value), float(ws.cell(x, y + 2).value)]

                elif group is None:
                    break
                else:
                    return {None: group}
        elif sh is None:
            break
        else:
            return {None: sh}
    return result


async def get_name_groups():
    url_get_groups = "https://api.moysklad.ru/api/remap/1.2/entity/productfolder"
    response = requests.get(url_get_groups, headers=token.headers)
    data = json.loads(response.text)

    groups = {}
    for group in data['rows']:
        if group['name'] not in groups:
            groups[group['name']] = [(group['id'], group['pathName'])]
        else:
            groups[group['name']].append((group['id'], group['pathName']))

    return groups


async def cm_start_get_plans(message: types.Message):
    if message.from_user.id in inf.get_admin_id():
        kb = ReplyKeyboardMarkup(resize_keyboard=True)
        kb.insert(datetime.now().strftime('%m.%Y')).row('Отмена')
        await bot.send_message(message.from_user.id, f'Введите месяц и год для плана продаж:', reply_markup=kb)
        await FSMGetPlans.date.set()
    else:
        try:
            point = None
            for ret in sqlite_db.cur.execute("SELECT now_user, point1 FROM reports_open WHERE report_close_id = ?", ['']).fetchall():
                if str(inf.get_user_id(ret[0])) == str(message.from_user.id):
                    point = ret[1]
            if point is None:
                await bot.send_message(message.from_user.id, f'Вы не открыли смену')
                return
            mes = await bot.send_message(message.from_user.id, f"Подождите, идёт загрузка...")
            text = await get_plans(point, datetime.now().strftime("%m.%Y"))
            await bot.delete_message(chat_id=message.from_user.id, message_id=mes.message_id)
            await bot.send_message(message.from_user.id, text)
        except Exception:
            await bot.send_message(message.from_user.id, f"Произошла ошибка, обратитесь к старшему сотруднику")


async def plans_get_date(message: types.Message,  state: FSMContext):

    try:
        date_ = datetime.strptime(message.text, "%m.%Y").strftime("%m.%Y")
    except Exception:
        await bot.send_message(message.from_user.id, f'Формат(ММ.ГГГГ)')
        return

    async with state.proxy() as data:
        data['date'] = date_

    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    for store in inf.get_name_shops():
        kb.insert(store)
    kb.row('Отмена')
    await bot.send_message(message.from_user.id, f'Выберите точку для выгрузки плана продаж', reply_markup=kb)
    await FSMGetPlans.next()


async def plans_get_point(message: types.Message,  state: FSMContext):
    try:
        async with state.proxy() as data:
            mes = await bot.send_message(message.from_user.id, f'Подождите, идёт загрузка...',
                                         reply_markup=types.ReplyKeyboardRemove())
            text = await get_plans(message.text, data['date'])

            await bot.delete_message(chat_id=message.from_user.id, message_id=mes.message_id)
            await bot.send_message(message.from_user.id, text, reply_markup=inf.kb(message.from_user.id))
    except Exception:
        await bot.send_message(message.from_user.id, f'Произошла ошибка, невозможно выгрузить планы. ',
                               reply_markup=inf.kb(message.from_user.id))
    await state.finish()


async def get_plans(point, date, count = False):
    month, year = date.split('.')
    sklad_id = ""
    for ret in sqlite_db.cur.execute('SELECT id_sklad FROM shops WHERE name_point = ?', [point]):
        sklad_id = ret[0]
    filter_sklad = f"retailStore=https://api.moysklad.ru/api/remap/1.2/entity/retailstore/{sklad_id}"

    date_from = f'{year}-{month}-01 08:00:00'
    date_to = f'{year}-{month}-{calendar.monthrange(int(year), int(month))[1]} 23:59:59'

    count_group = {}
    text = f"Планы продаж на {month}.{year}, на точку {point}:\n"

    for ret in sqlite_db.cur.execute(
            'SELECT group_name, value, group1, prem FROM plans_shop WHERE store_id = ? AND month = ? AND year = ?',
            [inf.get_id_point(point), month, year]):
        count_group[ret[2]] = [0, float(ret[1]), float(ret[3])]
        filter_group = f"productFolder=https://api.moysklad.ru/api/remap/1.2/entity/productfolder/{ret[2]}"
        url_get_plans = f"https://api.moysklad.ru/api/remap/1.2/report/profit/byproduct?filter={filter_sklad};" \
                        f"{filter_group}"

        for i in range(0, 1000):
            response = requests.get(url_get_plans, headers=token.headers, params={'offset': i * 1000,
                                                                                  'momentFrom': date_from,
                                                                                  'momentTo': date_to})
            data = json.loads(response.text)

            if len(data['rows']) == 0:
                break
            for product in data['rows']:
                count_group[ret[2]][0] += float(product['sellSum']) / 100
        text += f"{ret[0]}: {int(count_group[ret[2]][0])}/{int(float(ret[1]))} | Премия: {ret[3]}\n"
    if count:
        return count_group
    else:
        return text


async def get_pattern():
    wb = openpyxl.Workbook()
    ws = wb.active

    for i, store in enumerate(inf.get_name_shops()):
        ws.cell(1, i * 4 + 1).value = store
        ws.cell(2, i * 4 + 1).value = "Название группы"
        ws.cell(2, i * 4 + 2).value = "Значение"
        ws.cell(2, i * 4 + 3).value = "Премия"

    wb.save('pattern.xlsx')


async def take_prem_plans(message: types.Message):
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.insert(datetime.now().strftime('%m.%Y')).row('Отмена')
    await bot.send_message(message.from_user.id, f'Введите месяц и год для плана продаж:', reply_markup=kb)
    await FSMTakePlans.date.set()


async def take_data_plans(message: types.Message, state: FSMContext):
    try:
        date_ = datetime.strptime(message.text, "%m.%Y").strftime("%m.%Y")
    except Exception:
        await bot.send_message(message.from_user.id, f'Формат(ММ.ГГГГ)')
        return
    month, year = date_.split('.')
    text = f"Премии по выполнению плана продаж за {month}.{year}:\n"
    point_user, stores = {}, {}
    user_premies, store_premies = {}, {}

    for store in inf.get_name_shops():
        stores[store] = 0
        store_premies[store] = 0

        count_group = await get_plans(store, f"{month}.{year}", True)
        for group in count_group.keys():
            if count_group[group][0] > count_group[group][1]:
                store_premies[store] += count_group[group][2]

    for uid in inf.get_users_id():
        person = inf.get_name(uid)
        point_user[person], user_premies[person] = {}, 0
        for store in inf.get_name_shops():
            point_user[person][store] = 0

    for ret in sqlite_db.cur.execute('SELECT point1, person FROM menu WHERE month = ? AND year = ?',
                                     [month, year]):
        if ret[1] in list(point_user.keys()):
            point_user[ret[1]][ret[0]] += 1
        stores[ret[0]] += 1

    for store, count in stores.items():
        for person in point_user.keys():
            for store_user, count_day in point_user[person].items():
                if store_user == store:
                    if count != 0:
                        user_premies[person] += round(store_premies[store] * (count_day / count), 0)

    sqlite_db.cur.execute("DELETE FROM premies_plan WHERE month = ? AND year = ?", [month, year])
    sqlite_db.cur.execute("DELETE FROM premies_plan_point WHERE month = ? AND year = ?", [month, year])
    for user in user_premies.keys():
        text += f"{user}: {user_premies[user]} руб.\n"
        sqlite_db.cur.execute("INSERT INTO premies_plan VALUES (?,?,?,?)", [user, user_premies[user], month, year])
    for store in store_premies.keys():
        sqlite_db.cur.execute("INSERT INTO premies_plan_point VALUES (?,?,?,?)", [inf.get_id_point(store), store_premies[store], month, year])
    sqlite_db.base.commit()

    await bot.send_message(message.from_user.id, text, reply_markup=inf.kb(message.from_user.id))
    await state.finish()


async def take_now_plan(message: types.Message):
    try:
        with open(f'uploads/plans_of_sales.xlsx', 'rb') as file:
            await bot.send_document(message.from_user.id, file)
    except Exception:
        await bot.send_message(message.from_user.id, f"Файл не обнаружен")



def register_handlers_plans(dp: Dispatcher):
    dp.register_message_handler(cm_start_plans, Text(equals=f'📔Создать план продаж', ignore_case=True),state=None)
    dp.register_message_handler(plans_take_date, state=FSMPlans.date)
    dp.register_message_handler(plans_get, content_types=types.ContentTypes.DOCUMENT, state=FSMPlans.id)
    dp.register_message_handler(cm_start_get_plans, Text(equals=f'📔План продаж', ignore_case=True), state=None)
    dp.register_message_handler(plans_get_date, state=FSMGetPlans.date)
    dp.register_message_handler(plans_get_point, state=FSMGetPlans.point)
    dp.register_message_handler(take_prem_plans, Text(equals=f'📔Начислить премии по планам', ignore_case=True),
                                state=None)
    dp.register_message_handler(take_data_plans, state=FSMTakePlans.date)
    dp.register_message_handler(take_now_plan, Text(equals=f'📔Таблица с планом', ignore_case=True),
                                state=None)

