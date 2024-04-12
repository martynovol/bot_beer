import asyncio
import json
import traceback
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import openpyxl
import requests
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram import types, Dispatcher
from aiogram.dispatcher.filters import Text
from aiogram.types import CallbackQuery, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, InlineKeyboardMarkup, \
	InlineKeyboardButton
from aiogram.types import InputFile
import numpy as np
from create_bot import dp, bot
from database import sqlite_db
from PIL import Image
from keyboards import admin_kb, admin_cancel_kb, cassier_kb, mod_kb, kb_client
from id import token
import matplotlib.pyplot as plt
import os
from datetime import datetime
from datetime import date, timedelta
import calendar
from handlers import inf
from handlers import emoji_bot

import emoji


class FSMAdmin_report_month(StatesGroup):
    month1 = State()


class FSMAdmin_month_excel(StatesGroup):
    month = State()


#Отчёт за месяц
async def month_report(message: types.Message):
    admins = inf.get_admin_id()
    if message.from_user.id not in admins:
        await bot.send_message(message.from_user.id, 'Эта функция доступна только администраторам')
        return
    await FSMAdmin_report_month.month1.set()
    b11 = KeyboardButton(str(datetime.now().strftime('%m.%Y')))
    b12 = KeyboardButton('Назад')
    button_case_admin_report = ReplyKeyboardMarkup(resize_keyboard=True).add(b11).row(b12)
    await bot.send_message(message.from_user.id,
                           f'Впишите месяц и год за который нужно выгрузить общую выручку в формате ММ.ГГГГ{emoji_bot.em_report_load_date}:',
                           reply_markup=button_case_admin_report)


# Выгрузка прибыли за месяц
async def take_report_month(message: types.Message, state: FSMContext):
    bit = message.text.split('.')
    if len(bit) != 2 or (len(bit[0]) != 1 and len(bit[0]) != 2) or len(bit[1]) != 4 or not bit[0].isdigit() or not bit[1].isdigit():
        await message.reply(f'Вы ввели некорректное число, попробуйте снова, формат "ММ.ГГГГ"')
        return
    for mod_id in inf.get_mod_id():
        time = str(datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        await bot.send_message(mod_id, f'{inf.get_name(str(message.from_user.id))}, {message.from_user.id}, {time} получает выручку за месяц: {message.text}')
    def take_summ(month, year):
        if len(str(month)) == 1:
            month = '0' + str(month)
        if int(month) == 0:
            year -= 1
            month = 12
        data1 = [0, 0, 0, 0, 0, 0]
        shops = {}
        last_day = get_last_day(month,year)
        for ret in sqlite_db.cur.execute("SELECT * FROM menu WHERE month LIKE :month1 AND year LIKE :year1 "
                                         "ORDER BY year ASC, month ASC, day ASC",
                                         {'month1': str(month), 'year1': str(year)}).fetchall():
            if int(ret[3]) <= last_day:
                if len(ret[7].split(' ')) == 3:
                    beznal, sbp, qr = ret[7].split(' ')[0], ret[7].split(' ')[1], ret[7].split(' ')[2]
                elif len(ret[7].split(' ')) == 2:
                    beznal, sbp, qr = ret[7].split(' ')[0], ret[7].split(' ')[1], '0.0'
                else:
                    beznal, sbp, qr = ret[7], '0.0', '0.0'
                a = [ret[6], beznal, ret[8], ret[9], sbp, qr]
                if ret[1] not in shops:
                    shops[ret[1]] = [0, 0, 0, 0, 0, 0]
                for i in range(0, 6):
                    data1[i] += round(float(a[i]), 2)
                    shops[ret[1]][i] += round(float(a[i]), 2)
        if data1[0] == 0 and data1[1] == 0 and data1[2] == 0:
            return None
        return data1[0], data1[1], data1[2], data1[3], data1[4], shops, data1[5]
    if take_summ(int(bit[0]), int(bit[1])) == None:
        await state.finish()
        await bot.send_message(message.from_user.id, f'За этот период выгрузка не найдена', reply_markup=inf.kb(message.from_user.id))
        return 
    cash1, non_cash1, transfers, summ, sbp, shops, qr = take_summ(int(bit[0]), int(bit[1]))
        #await bot.send_message(message.from_user.id, text='^^^', reply_markup=InlineKeyboardMarkup().add(
        #    InlineKeyboardButton(f'Удалить за {ret[2]}', callback_data=f'del {ret[2]}')))
    s = {}
    counter = {}
    prognoz = {}
    last_day = get_last_day(bit[0], bit[1])
    for ret in sqlite_db.cur.execute('SELECT * FROM menu ORDER BY year ASC, month ASC, day ASC'):
        date1 = f'{ret[4]} {ret[5]}'
        if ret[1] not in s:
            s[ret[1]] = [float(ret[9]), [date1]]
        else:
            now_date = str(str(date.today()).split('-')[1]) +  ' ' + str(date.today()).split('-')[0]
            if date1 != now_date and int(ret[3]) <= last_day:
                if date1 not in s[ret[1]][1]:
                    s[ret[1]][1].append(date1)
                s[ret[1]][0] += float(ret[9])
            now_date = str(date.today()).split('-')[1] +  ' ' + str(date.today()).split('-')[0] #Выручки все здесь
    em1 = emoji.emojize(":red_triangle_pointed_up:")
    em2 = emoji.emojize(":red_triangle_pointed_down:")
    for i in range(0, len(shops)):
        name_shop = list(shops.keys())[i]
        average = round(s[name_shop][0] / len(s[name_shop][1]), 2)
        if average == 0:
            average = 1
        raz2 = round((shops[name_shop][3] - average)/average*100, 2)
        keyboard = InlineKeyboardMarkup().add(InlineKeyboardButton('Анализ', callback_data=f'pgraph {inf.get_id_point(name_shop)} {bit[0]} {bit[1]}'))
        if raz2 > 0:
            await bot.send_message(message.from_user.id, f'Выручка за {bit[0]}.{bit[1]} с магазина \'{name_shop}\':\nНаличными: {round(shops[name_shop][0], 2)} руб.\nТерминал: {round(shops[name_shop][1], 2)} руб.\nПереводами: {round(shops[name_shop][2], 2)} руб.\nСБП: {round(shops[name_shop][4], 2)}\nQR-Кодом: {round(shops[name_shop][5], 2)}\n----------------------------------------------\nСредняя выручка магазина до {last_day} числа \'{name_shop}\': {average} руб.\nВыручка с магазина \'{name_shop}\' в этом месяце выше средней на {round(((shops[name_shop][3] - average)/average)*100, 1)}% {em1}\n--------------------------------------------\nОбщая выручка: {round(shops[name_shop][3], 2)} руб.', reply_markup=keyboard)
        else:
            await bot.send_message(message.from_user.id, f'Выручка за {bit[0]}.{bit[1]} с магазина \'{name_shop}\':\nНаличными: {round(shops[name_shop][0], 2)} руб.\nТерминал: {round(shops[name_shop][1], 2)} руб.\nПереводами: {round(shops[name_shop][2], 2)} руб.\nСБП: {round(shops[name_shop][4], 2)}\nQR-Кодом: {round(shops[name_shop][5], 2)}\n----------------------------------------------\nСредняя выручка магазина до {last_day} числа \'{name_shop}\': {average} руб.\nВыручка с магазина \'{name_shop}\' в этом месяце ниже средней на {round(((shops[name_shop][3] - average)/average)*100, 1)}% {em2}\n--------------------------------------------\nОбщая выручка: {round(shops[name_shop][3], 2)} руб.', reply_markup=keyboard)
    if take_summ(int(bit[0]) - 1, int(bit[1])) == None:
        await bot.send_message(message.from_user.id, f'Выручка за {bit[0]}.{bit[1]} со всех магазинов:\nНаличными: {round(cash1, 2)} руб.\nТерминал: {round(non_cash1, 2)} руб.\nПереводами: {round(transfers,2)} руб.\n--------------------------------------------\nОбщая выручка: {round(summ, 2)} руб.')
    else:
        if datetime.strftime(datetime.now(), "%m.%Y") == f"{bit[0]}.{bit[1]}":
            cash2, non_cash2, transfers2, summ2, sbp2, shops2, qr2 = take_summ_last_month(int(bit[0]) - 1, int(bit[1]))
        else:
            cash2, non_cash2, transfers2, summ2, sbp2, shops2, qr2 = take_summ(int(bit[0]) - 1, int(bit[1]))
        total = 0
        total_count = []
        for i in range(0, len(shops)):
            name_shop = list(shops.keys())[i]
            total += s[name_shop][0]
            for j in s[name_shop][1]:
                if j not in total_count:
                    total_count.append(j)
        average = round(total / len(total_count), 2)
        if average == 0:
            average = 1
        raz1 = round(((summ - summ2)/summ2)*100, 1)
        raz2 = round(((summ - average)/average)*100, 1)
        keyboard = InlineKeyboardMarkup().add(InlineKeyboardButton('Анализ',
                                                                   callback_data=f'agraph {bit[0]} {bit[1]}'))
        if raz1 > 0:
            if raz2 > 0:
                await bot.send_message(message.from_user.id, f'Выручка за {bit[0]}.{bit[1]} со всех магазинов:\nНаличными: {round(cash1, 2)} руб.\nТерминал: {round(non_cash1, 2)} руб.\nСБП: {round(sbp, 2)}\nQR-Кодом: {round(qr, 2)}\nПереводами: {round(transfers,2)} руб.\n--------------------------------------------\nСредняя выручка: {average} руб.\nВыручка в этом месяце выше средней на {round(((summ - average)/average)*100, 1)}% {em1}\nВыручка в этом месяце выросла относительно прошлого на {round(((summ - summ2)/summ2)*100, 1)}% {em1}\n--------------------------------------------\nОбщая выручка: {round(summ, 2)} руб.', reply_markup=keyboard)
            else:
                await bot.send_message(message.from_user.id, f'Выручка за {bit[0]}.{bit[1]} со всех магазинов:\nНаличными: {round(cash1, 2)} руб.\nТерминал: {round(non_cash1, 2)} руб.\nСБП: {round(sbp, 2)}\nQR-Кодом: {round(qr, 2)}\nПереводами: {round(transfers,2)} руб.\n--------------------------------------------\nСредняя выручка: {average} руб.\nВыручка в этом месяце ниже средней на {round(((summ - average)/average)*100, 1)}% {em2}\nВыручка в этом месяце выросла относительно прошлого на {round(((summ - summ2)/summ2)*100, 1)}% {em1}\n--------------------------------------------\nОбщая выручка: {round(summ, 2)} руб.', reply_markup=keyboard)
        else:
            if raz2 > 0:
                await bot.send_message(message.from_user.id, f'Выручка за {bit[0]}.{bit[1]} со всех магазинов:\nНаличными: {round(cash1, 2)} руб.\nТерминал: {round(non_cash1, 2)} руб.\nСБП: {round(sbp, 2)}\nQR-Кодом: {round(qr, 2)}\nПереводами: {round(transfers,2)} руб.\n--------------------------------------------\nСредняя выручка: {average} руб.\nВыручка в этом месяце выше средней на {round(((summ - average)/average)*100, 1)}% {em1}\nВыручка в этом месяце сократилась относительно прошлого на {round(((summ - summ2)/summ2)*100, 1)}% {em2}\n--------------------------------------------\nОбщая выручка: {round(summ, 2)} руб.', reply_markup=keyboard)
            else:
                await bot.send_message(message.from_user.id, f'Выручка за {bit[0]}.{bit[1]} со всех магазинов:\nНаличными: {round(cash1, 2)} руб.\nТерминал: {round(non_cash1, 2)} руб.\nСБП: {round(sbp, 2)}\nQR-Кодом: {round(qr, 2)}\nПереводами: {round(transfers,2)} руб.\n--------------------------------------------\nСредняя выручка: {average} руб.\nВыручка в этом месяце ниже средней на {round(((summ - average)/average)*100, 1)}% {em2}\nВыручка в этом месяце сократилась относительно прошлого на {round(((summ - summ2)/summ2)*100, 1)}% {em2}\n--------------------------------------------\nОбщая выручка: {round(summ, 2)} руб.', reply_markup=keyboard)
    await state.finish()
    await bot.send_message(message.from_user.id, f'Выгрузка завершена', reply_markup=inf.kb(message.from_user.id))


# Кнопка 'назад' при отправке репортов
async def cancel_handler_rep(message: types.Message, state: FSMContext):
    for mod_id in inf.get_mod_id():
        time = str(datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        await bot.send_message(mod_id, f'{inf.get_name(str(message.from_user.id))}, {message.from_user.id}, {time}: {message.text}')
    kb = mod_kb.button_case_mod if message.from_user.id in inf.get_mod_id() else admin_kb.button_case_admin
    current_state = await state.get_state()
    if current_state is None:
        await message.reply('Вы вернулись назад', reply_markup=kb)
        return
    await state.finish()
    await message.reply('Вы вернулись назад', reply_markup=kb)


def get_last_day(month, year):
    last_day = 0
    for ret in sqlite_db.cur.execute("SELECT * FROM menu WHERE month LIKE ? AND year LIKE ?"
                                     "ORDER BY year ASC, month ASC, day ASC", [month, year]).fetchall():
        last_day = int(ret[3])
    return last_day


def take_summ_last_month(month, year):
        if len(str(month)) == 1:
            month = '0' + str(month)
        if int(month) == 0:
            year -= 1
            month = 12
        data1 = [0, 0, 0, 0, 0, 0]
        shops = {}
        for ret in sqlite_db.cur.execute("SELECT * FROM menu WHERE month LIKE :month1 AND year LIKE :year1", {'month1':str(month),'year1': str(year)}).fetchall():
            if int(int(ret[2].split('.')[0])) <= int(datetime.strftime(datetime.now(), "%d")):
                if len(ret[7].split(' ')) == 3:
                    beznal, sbp, qr = ret[7].split(' ')[0], ret[7].split(' ')[1], ret[7].split(' ')[2]
                elif len(ret[7].split(' ')) == 2:
                    beznal, sbp, qr = ret[7].split(' ')[0], ret[7].split(' ')[1], '0.0'
                else:
                    beznal, sbp, qr = ret[7], '0.0', '0.0'
                a = [ret[6], beznal, ret[8], ret[9], sbp, qr]
                if ret[1] not in shops:
                    shops[ret[1]] = [0, 0, 0, 0, 0, 0]
                for i in range(0, 6):
                    data1[i] += round(float(a[i]), 2)
                    shops[ret[1]][i] += round(float(a[i]), 2)
        if data1[0] == 0 and data1[1] == 0 and data1[2] == 0:
            return None
        return data1[0], data1[1], data1[2],data1[3], data1[4], shops, data1[5]


async def load_graph(callback_query: types.CallbackQuery):
    uid = callback_query.from_user.id
    mes = await bot.send_message(uid, "Подождите анализируется")
    point = inf.pt_name(callback_query.data.split(' ')[1])
    month = callback_query.data.split(' ')[2]
    year = callback_query.data.split(' ')[3]

    list_days, list_persons, days_count, total_days = {}, {}, {}, {}
    total_days = [0, 0, 0, 0]
    for ret in sqlite_db.cur.execute("SELECT day, total, person, cash, non_cash, transfers FROM menu WHERE month"
                                     " LIKE :month1 AND year LIKE :year1 AND point1 LIKE :point1",
                                     {'month1': month, 'year1': year, 'point1': point}).fetchall():
        list_days[str(int(ret[0]))] = float(ret[1])
        total_days[0] += float(ret[3])
        total_days[1] += float(ret[4].split(' ')[0])
        if len(ret[4].split(' ')) == 2:
            total_days[2] += float(ret[4].split(' ')[1])
        total_days[3] += float(ret[5])


        if ret[2] in list_persons:
            list_persons[ret[2]] += float(ret[1])
            days_count[ret[2]] += 1
        else:
            list_persons[ret[2]] = float(ret[1])
            days_count[ret[2]] = 1

    list_days = dict(sorted(list_days.items(), key=lambda x: int(x[0])))
    for key in list(list_persons.keys()):
        list_persons[f'{key}\n{days_count[key]} дн.'] = list_persons.pop(key)

    label_money = ['Наличными', 'Безналичными', 'СБП', 'Переводами']
    for i in range(len(total_days)):
        try:
            if total_days[i] == 0:
                total_days.pop(i)
                label_money.pop(i)
                i -= 1
        except Exception:
            break

    def func(pct, allvalues):
        absolute = int(pct / 100. * np.sum(allvalues))
        return "{:.1f}%\n({:d} руб.)".format(pct, absolute)

    plt.figure(figsize=(9, 9), facecolor="seashell")
    plt.bar(list_days.keys(), list_days.values())
    plt.title(f'----\nВыручка по дням на точке {point} {month}.{year}\nВсего:{round(sum(list_days.values()),0)} руб.\n'
              f'Средняя:{round(sum(list_days.values())/len(list_days.values()), 0)} руб.')
    plt.xlabel('День')
    plt.ylabel('Выручка')
    plt.xticks(rotation=90, ha='center')
    plt.tight_layout()
    plt.savefig(f'photo{str(uid)}{1}.png', dpi=400)
    plt.close()

    plt.figure(figsize=(9, 9), facecolor="seashell")
    plt.pie(list_persons.values(), labels=list_persons.keys(),
            autopct=lambda pct: func(pct, list(list_persons.values())),
            wedgeprops={'lw': 1, 'ls': '--', 'edgecolor': "k"})
    plt.title(f'Продажи на точке {point} {month}.{year}')
    plt.tight_layout()
    plt.savefig(f'photo{str(uid)}{2}.png', dpi=400)
    plt.close()

    plt.figure(figsize=(9, 9), facecolor="seashell")
    plt.pie(total_days, labels=label_money,
            autopct='%1.1f%%',
            wedgeprops={'lw': 1, 'ls': '--', 'edgecolor': "k"})
    plt.title(f'Продажи на точке {point} {month}.{year}')
    plt.tight_layout()
    plt.savefig(f'photo{str(uid)}{3}.png', dpi=400)
    plt.close()



    media = types.MediaGroup()
    media.attach_photo(types.InputFile(f'photo{str(uid)}{1}.png'))
    media.attach_photo(types.InputFile(f'photo{str(uid)}{2}.png'))
    media.attach_photo(types.InputFile(f'photo{str(uid)}{3}.png'))

    await bot.send_media_group(uid, media=media)

    os.remove(f'photo{str(uid)}{1}.png')
    os.remove(f'photo{str(uid)}{2}.png')
    os.remove(f'photo{str(uid)}{3}.png')
    await bot.delete_message(callback_query.from_user.id, mes.message_id)


async def load_graph_total(callback_query: types.CallbackQuery):
    uid = callback_query.from_user.id
    mes = await bot.send_message(uid, "Подождите анализируется")
    month = callback_query.data.split(' ')[1]
    year = callback_query.data.split(' ')[2]
    total_money = get_total(month, year)
    total_shops, color_rectangle, list_persons = {}, {}, {}
    for ret in sqlite_db.cur.execute("SELECT point1, day, total, person, cash, non_cash, transfers FROM menu WHERE month"
                                     " LIKE :month1 AND year LIKE :year1",
                                     {'month1': month, 'year1': year}).fetchall():
        if ret[3] in list_persons:
            list_persons[ret[3]][0] += float(ret[2])
            list_persons[ret[3]][1] += 1
        else:
            list_persons[ret[3]] = [float(ret[2]), 1]

        if ret[0] in total_shops:
            total_shops[ret[0]] += float(ret[2])
        else:
            total_shops[ret[0]] = float(ret[2])

    for key in list_persons.keys():
        list_persons[key] = round(list_persons[key][0] / list_persons[key][1], 1)
    sorted_tuple = sorted(list_persons.items(), key=lambda x: x[1])
    list_persons = dict(list(reversed(sorted_tuple)))
    sorted_tuple = sorted(total_shops.items(), key=lambda x: x[1])
    total_shops = dict(list(reversed(sorted_tuple)))
    for key in total_shops.keys():
        color_rectangle[key] = np.random.rand(1, 3)
    plt.figure(figsize=(9, 9), facecolor="seashell")
    handles, labels = [], []

    for key, value in total_shops.items():
        handle = plt.barh(key, value, color=color_rectangle[key])
        handles.append(handle)
        labels.append(key)

    plt.title(f'Выручка на точках {month}.{year}\nВсего: {round(sum(total_shops.values()), 0)} руб.\n'
              f'Средняя: {round(sum(total_shops.values()) / len(total_shops.values()), 0)} руб.')
    plt.ylabel('Точка')
    plt.xlabel('Выручка')
    plt.xticks(rotation=90, ha='center')
    for key, value in total_shops.items():
        plt.text(value, key, str(value))
    plt.tight_layout()
    plt.savefig(f'photo{str(uid)}{11}.png', dpi=400)
    plt.close()
    #ПРОДАВЦЫ
    plt.figure(figsize=(9, 9), facecolor="seashell")
    for key in list_persons.keys():
        color_rectangle[key] = np.random.rand(1, 3)

    handles, labels = [], []
    for key, value in list_persons.items():
        handle = plt.barh(key, value, color=color_rectangle[key])
        handles.append(handle)
        labels.append(key)

    plt.title(f'Средние продажи по сотрудникам {month}.{year}\n')
    plt.ylabel('Продавец')
    plt.xlabel('Среднее продаж')
    plt.xticks(rotation=90, ha='center')
    for key, value in list_persons.items():
        plt.text(value, key, str(value))# ha='center', va='bottom', rotation=90)
    plt.tight_layout()
    plt.savefig(f'photo{str(uid)}{12}.png', dpi=400)
    plt.close()
    #Общая выручка
    fig = plt.figure(figsize=(9, 9), facecolor="seashell")
    plt.bar(total_money.keys(), total_money.values())
    plt.title(f'----\nВыручка по дням на всех точках {month}.{year}\nВсего:{round(sum(total_money.values()), 0)} руб.\n'
              f'Средняя:{round(sum(total_money.values()) / len(total_money.values()), 0)} руб.')
    plt.xlabel('Месяц')
    plt.ylabel('Выручка')
    plt.xticks(rotation=90, ha='center')
    axis_data = fig.gca()
    axis_data.yaxis.get_major_formatter().set_scientific(False)
    axis_data.yaxis.get_major_formatter().set_useOffset(False)
    for key, value in total_money.items():
        plt.text(key, value, str(round(value, 0)), ha='center', va='bottom')
    plt.tight_layout()
    plt.savefig(f'photo{str(uid)}{13}.png', dpi=400)
    plt.close()

    media = types.MediaGroup()
    media.attach_photo(types.InputFile(f'photo{str(uid)}{11}.png'))
    media.attach_photo(types.InputFile(f'photo{str(uid)}{12}.png'))
    media.attach_photo(types.InputFile(f'photo{str(uid)}{13}.png'))
    await bot.send_media_group(uid, media=media)

    os.remove(f'photo{str(uid)}{11}.png')
    os.remove(f'photo{str(uid)}{12}.png')
    os.remove(f'photo{str(uid)}{13}.png')
    await bot.delete_message(callback_query.from_user.id, mes.message_id)



def get_total(month, year):
    total_money = {}
    for ret in sqlite_db.cur.execute('SELECT * FROM menu WHERE year LIKE ? OR year LIKE ?', [year, str(int(year) - 1)]):
        if (ret[5] != year and ret[4] >= month) or (ret[5] == year and ret[4] <= month):
            date1 = f'{ret[4]}.{ret[5]}'
            if date1 not in total_money:
                total_money[date1] = float(ret[9])
            else:
                total_money[date1] += float(ret[9])
    return total_money


async def month_report_excel(message: types.Message):
    admins = inf.get_mod_id()
    if message.from_user.id not in admins:
        await bot.send_message(message.from_user.id, 'Эта функция доступна только администраторам')
        return
    month = datetime.now().strftime("%m.%Y")
    kb = ReplyKeyboardMarkup(resize_keyboard=True).add(month).add('Отмена')
    await bot.send_message(message.from_user.id, f'Выберите месяц и год в формате (ММ.ГГГГ)', reply_markup=kb)
    await FSMAdmin_month_excel.month.set()


def get_salary_storage(month, year):
    list_privileges = {"manager_report": 2500, "operator_report": 1600, "driver_report": 1600, "storager_report": 2050}
    total_sum = 0
    for privilegie, salary in list_privileges.items():
        for ret in sqlite_db.cur.execute(f"SELECT * FROM {privilegie} WHERE month = ? AND year = ?", [month, year]).fetchall():
            total_sum += salary
    return total_sum


async def report_excel_1(message: types.Message, state: FSMContext):
    try:
        x = await bot.send_message(message.from_user.id, f'Подождите, идёт загрузка...')
        await sqlite_db.info_admin(message.from_user.id, "Начал грузить ОПУ")
        await asyncio.sleep(2)
        month, year = message.text.split('.')
        stores = inf.get_name_shops()
        a_revenue, a_defects, a_salaries, a_month_revenue, a_add_expenses, a_plans, a_revision =\
            {"Наличные": 0, "Безналичные": 0, "СБП": 0, "Переводы": 0, "QR-Код": 0}, 0, 0, 0, {}, 0, 0
        
        fines, a_point = get_fine_point(month, year)
        sal_storage = 0

        salary_point = {}
        for store in stores:
            salary_point[store] = 0
        
        totals = 0
        for user in inf.get_users():
            dict_day = inf.get_user_salary(user[1], month, year, True)
            if len(dict_day) == 2:
                dict_day = dict_day[1]
                for day, value in dict_day.items():
                    salary_point[value[4]] += (float(value[0]) + float(value[1]))

        sal_storage = get_salary_storage(month, year)

        for i, store in enumerate(stores):
            if store == "Склад":
                continue
            x = await bot.edit_message_text(chat_id=message.from_user.id, message_id=x.message_id,
                                            text=f'Загружается... {store} [{i + 1}/{len(stores)}]')
            revenue, defects, salaries, plans, revision = {"Наличные": 0, "Безналичные": 0, "СБП": 0, "Переводы": 0, "QR-Код": 0}, 0, 0, 0, 0
            for ret in sqlite_db.cur.execute("SELECT cash, non_cash, transfers, total, person, date1 FROM menu "
                                            "WHERE month LIKE ? AND year LIKE ? AND point1 LIKE ?", [month, year, store]).fetchall():
                revenue["Наличные"] += float(ret[0])
                revenue["Безналичные"] += float(ret[1].split(' ')[0])
                revenue["СБП"] += float(ret[1].split(' ')[1])
                revenue["Переводы"] += float(ret[2])
                if len(ret[1].split(' ')) == 3:
                    revenue["QR-Код"] += float(ret[1].split(' ')[2])
                    a_revenue["QR-Код"] += float(ret[1].split(' ')[2])
                else:
                    revenue["QR-Код"] += 0
                    a_revenue["QR-Код"] += 0
                a_revenue["Наличные"] += float(ret[0])
                a_revenue["Безналичные"] += float(ret[1].split(' ')[0])
                a_revenue["СБП"] += float(ret[1].split(' ')[1])
                a_revenue["Переводы"] += float(ret[2])

            salaries += salary_point[store]
            a_salaries += salary_point[store]

            for ret in sqlite_db.cur.execute("SELECT premie FROM premies_plan_point WHERE point = ? AND month = ? AND year = ?", [inf.get_id_point(store), month, year]):
                plans += float(ret[0])
                a_plans += float(ret[0])
            

            for ret in sqlite_db.cur.execute("SELECT price, type FROM defects WHERE point_name LIKE ? AND month = ? AND year = ?",
                                            [inf.get_id_point(store), month, year]):
                if ret[1] == "Брак":                         
                    defects += float(ret[0]) / 100
                    a_defects += float(ret[0]) / 100
            
            for ret in sqlite_db.cur.execute("SELECT value FROM revision WHERE point LIKE ? AND month = ? AND year = ?",
                                            [inf.get_id_point(store), month, year]).fetchall():
                revision += float(ret[0])
                a_revision += float(ret[0])

            month_revenue = get_month_revenue(month, year, store)
            add_expense = get_addit_expense(store)
            for key, value in add_expense.items():
                if key not in a_add_expenses:
                    a_add_expenses[key] = value
                else:
                    a_add_expenses[key] += value
            a_month_revenue += month_revenue
            fine = 0
            if store in fines:
                fine = fines[store]
            await in_file(revenue, salaries, defects, month_revenue, month, year, add_expense, message.from_user.id, plans, fine, revision, sal_storage,store)
        await in_file(a_revenue, a_salaries, a_defects, a_month_revenue, month, year, a_add_expenses, message.from_user.id, a_plans, a_point, a_revision, sal_storage)
        await bot.delete_message(chat_id=message.from_user.id, message_id=x.message_id)
        await bot.send_message(message.from_user.id, f'Загрузка завершена', reply_markup=inf.kb(message.from_user.id))
        with open(f'ОПУ_{message.from_user.id}_{month}_{year}.xlsx', 'rb') as file:
            await bot.send_document(message.from_user.id, file)
        try:
            os.remove(f'ОПУ_{message.from_user.id}_{month}_{year}.xlsx')
        except Exception:
            pass
        await state.finish()
    except Exception as e:
        traceback.print_exc()
        await bot.send_message(message.from_user.id, f"Произошла ошибка, обратитесь к программисту\n\n{str(e)[0:1024]}", reply_markup=inf.kb(message.from_user.id))
        await state.finish()

async def in_file(revenue, salaries, defects, month_revenue, month, year, add_expense, uid, plans, fines, revision, sal_storage,store_name="Общая"):
    try:
        wb = openpyxl.load_workbook(filename=f'ОПУ_{uid}_{month}_{year}.xlsx')
    except Exception:
        wb = openpyxl.Workbook()
    if store_name not in wb.sheetnames:
        wb.create_sheet(store_name)
    wb.active = wb[store_name]
    sheet = wb[store_name]
    ws = wb.active
    ws.merge_cells("A1:B1")
    ws.merge_cells("A9:B9")
    ws.merge_cells("E9:F9")
    ws.merge_cells("E1:F1")
    ws['A1'].font, ws['A9'].font, ws['E1'].font, ws['E9'].font = \
        Font(bold=True), Font(bold=True), Font(bold=True), Font(bold=True)
    ws['A1'].fill, ws['A9'].fill, ws['E1'].fill, ws['E9'].fill = \
        PatternFill("solid", fgColor="B5B8B1"), PatternFill("solid", fgColor="B5B8B1"), PatternFill("solid", fgColor="B5B8B1"), PatternFill("solid", fgColor="B5B8B1")
    ws[f'A1'].value = "Выручка"
    ws[f'E1'].value = "Рентабельность"
    i = 2
    for key, value in revenue.items():
        sheet[f'A{i}'], sheet[f'B{i}'] = key + ": ", value
        i += 1
    all_revenue = sum(list(revenue.values()))
    sheet[f'A{i}'], sheet[f'B{i}'] = "Итого:", all_revenue
    ws[f'A{i + 2}'].value = "Расходы"
    ws[f'E{i + 2}'].value = "Отдельно от ФОТ"
    i += 3
    all_another = {"Штрафы": fines, "Ревизии": revision}
    if store_name == "Общая":
        all_expenses = {"ФОТ Продавцы": salaries, "ФОТ Склад+Упр": sal_storage,"План продаж": plans, "Браки": defects, "СБП 0,8%": revenue["СБП"] * 0.008, "QR-Код 2%": revenue["QR-Код"] * 0.02, "Эквайринг 2 %": revenue["Безналичные"] * 0.02}
    else:
        all_expenses = {"ФОТ Продавцы": salaries, "План продаж": plans, "Браки": defects, "СБП 0,8%": revenue["СБП"] * 0.008, "QR-Код 2%": revenue["QR-Код"] * 0.02, "Эквайринг 2 %": revenue["Безналичные"] * 0.02}
    sheet['E2'], sheet['F2'] = "Выручка: ", all_revenue
    sheet['E3'], sheet['F3'] = "Расходы: ", f"=SUM(B{i}:B1000)"
    sheet['E4'], sheet['F4'] = "Прибыль: ", f"={month_revenue} - F3"
    sheet[f'A{i}'] = "Расходы:"

    for key, value in all_expenses.items():
        sheet[f'A{i}'], sheet[f'B{i}'] = key, value
        i += 1

    for key, value in add_expense.items():
        sheet[f'A{i}'], sheet[f'B{i}'] = key, value
        i += 1

    i = 10
    for key, value in all_another.items():
        sheet[f'E{i}'], sheet[f'F{i}'] = key, value
        i += 1

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save(f'ОПУ_{uid}_{month}_{year}.xlsx')


def get_addit_expense(store_name):
    expenses = {}
    for ret in sqlite_db.cur.execute('SELECT name, value FROM categories_expenses WHERE point=?', [inf.get_id_point(store_name)]).fetchall():
        expenses[ret[0]] = float(ret[1])
    for ret in sqlite_db.cur.execute('SELECT name, value FROM categories_expenses WHERE point="All"').fetchall():
        if ret[0] not in list(expenses.keys()):
            expenses[ret[0]] = float(ret[1])
    return expenses


def get_month_revenue(month, year, store):
    last_day_month, result = calendar.monthrange(int(year), int(month))[1], 0
    start_day = f"{year}-{month}-01 00:00:00"
    finish_day = f"{year}-{month}-{last_day_month} 23:59:00"
    id_store = inf.get_id_sklad_point(store)
    for i in range(0, 100):
        url = "https://api.moysklad.ru/api/remap/1.2/report/profit/byproduct?filter=retailStore=" \
              "https://api.moysklad.ru/api/remap/1.2/entity/retailstore/" + id_store 
        response = requests.get(url, headers=token.headers, params={"momentFrom": start_day, "momentTo": finish_day,
                                                                    "limit": 1000, "offset": i * 1000})
        data = json.loads(response.text)
        if 'rows' not in data.keys():
            break
        if len(data['rows']) == 0:
            break
        for product in data['rows']:
            result += product['profit'] / 100
    return result



def get_fine_point(month, year):
    a_point = 0

    stores = {}
    for store in inf.get_name_shops():
        stores[store] = 0

    for ret in sqlite_db.cur.execute("SELECT person, date1, fine FROM fine WHERE month = ? AND year = ?", [month, year]).fetchall():
        person, date_report = ret[0], ret[1]
        point = ""

        for jet in sqlite_db.cur.execute("SELECT point1 FROM menu WHERE person = ? AND date1 = ?", [person, date_report]):
            point = jet[0]

        a_point += float(ret[2])
        if point != "":
            stores[jet[0]] += float(ret[2])
    
    return stores, a_point


def register_handlers_month_report(dp: Dispatcher):
    dp.register_callback_query_handler(load_graph, lambda x: x.data and x.data.startswith('pgraph '))
    dp.register_callback_query_handler(load_graph_total, lambda x: x.data and x.data.startswith('agraph '))
    dp.register_message_handler(cancel_handler_rep, state="*", commands='Назад')
    dp.register_message_handler(cancel_handler_rep, Text(equals='Назад', ignore_case=True), state="*")
    dp.register_message_handler(month_report, Text(equals=f'{emoji_bot.em_report_for_month}Выгрузить выручку за месяц', ignore_case=True))
    dp.register_message_handler(month_report_excel, Text(equals=f'Сформировать ОПУ', ignore_case=True))
    dp.register_message_handler(report_excel_1, state=FSMAdmin_month_excel.month)
    dp.register_message_handler(take_report_month, state=FSMAdmin_report_month.month1)